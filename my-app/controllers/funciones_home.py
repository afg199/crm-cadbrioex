
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file

#Almacena en la base de datos el registro de un nuevo  CLIENTE
def procesar_form_empleado(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO clientes (nombre, apellido, nif, provincia, poblacion, direccion, codigo_postal, telefono, email) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_cliente'], 
                            dataForm['apellido_cliente'],
                            dataForm['nif_cliente'], 
                            dataForm['provincia_cliente'], 
                            dataForm['poblacion_cliente'], 
                            dataForm['direccion_cliente'],
                            dataForm['cp_cliente'], 
                            dataForm['telefono_cliente'], 
                            dataForm['email_cliente'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'

#Almacena en la base de datos el registro de un nuevo EQUIPO de un cliente
def procesar_form_equipo(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO equipos (id_Clientes, marca, modelo, n_serie, garantia, meses_garantia, compra_biomex, fecha_compra, comentario) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (dataForm['id_cliente'], 
                            dataForm['marca_equipo'],
                            dataForm['modelo_equipo'], 
                            dataForm['ns_equipo'], 
                            dataForm['garantia_equipo'], 
                            dataForm['mesesgarantia_equipo'],
                            dataForm['compra_equipo'], 
                            dataForm['fecha_equipo'], 
                            dataForm['comentario_equipo'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_equipo: {str(e)}'

#Almacena en la base de datos el registro de un nuevo SERVICIO de un cliente
def procesar_form_servicio(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO servicios (id_Equipo, tipo_servicio, fecha_servicio, trabajador, horas, comentario) VALUES (%s, %s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (dataForm['id_equipo'], 
                            dataForm['tipo_servicio'],
                            dataForm['fecha_equipo'], 
                            dataForm['trabajador_servicio'], 
                            dataForm['horas_servicio'], 
                            dataForm['comentario_equipo'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_servicio: {str(e)}'

"""def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []"""


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_Clientes,
                        e.nombre, 
                        e.apellido,
                        e.nif,
                        e.poblacion,   
                        e.telefono
                    FROM clientes AS e
                    ORDER BY e.id_Clientes DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None

# Lista de EQUIPOS de un Cliente
def sql_lista_equipo_clienteBD(idCliente):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_Equipos,
                        e.id_Clientes,
                        e.marca,
                        e.modelo,
                        e.n_serie,
                        CASE
                            WHEN e.garantia = 1 THEN 'SI'
                            ELSE 'NO'
                        END AS garantia,
                        e.meses_garantia,
                        CASE
                            WHEN e.compra_biomex = 1 THEN 'SI'
                            ELSE 'NO'
                        END AS compra_biomex,
                        e.fecha_compra,
                        e.comentario
                    FROM equipos AS e
                    WHERE id_Clientes = %s
                    ORDER BY e.id_Equipos DESC
                    """)
                cursor.execute(querySQL, (idCliente,))
                equiposBD = cursor.fetchall()
        return equiposBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_equipo_clienteBD: {e}")
        return None
    
# Lista de SERVICIOS realizado en un equipo
def sql_lista_equipo_servicioBD(idEquipo):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_Servicios,
                        e.id_Equipo,
                        CASE
                            WHEN e.tipo_servicio = 1 THEN 'PUESTA EN MARCHA'
                            WHEN e.tipo_servicio = 2 THEN 'LIMPIEZA'
                            WHEN e.tipo_servicio = 3 THEN 'REPARACIÓN'
                            WHEN e.tipo_servicio = 4 THEN 'OTRO SERVICIO'
                            ELSE 'NO ESPECIFICADO'
                        END AS tipo_servicio,
                        e.fecha_servicio,
                        e.trabajador,
                        e.horas,
                        e.comentario
                    FROM servicios AS e
                    WHERE id_Equipo = %s
                    ORDER BY e.tipo_servicio ASC
                    """)
                cursor.execute(querySQL, (idEquipo,))
                equiposBD = cursor.fetchall()
        return equiposBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_equipo_servicioBD: {e}")
        return None

# Detalles del cliente
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_Clientes,
                        e.nombre, 
                        e.apellido,
                        e.nif,
                        e.provincia,
                        e.poblacion,
                        e.direccion,
                        e.codigo_postal,
                        e.telefono,
                        e.email,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM clientes AS e
                    WHERE id_Clientes =%s
                    ORDER BY e.id_Clientes DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None

# Detalles del EQUIPOS de un Cliente
def sql_detalles_equipoBD(idEquipo):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_Equipos,
                        e.id_Clientes,
                        e.marca,
                        e.modelo,
                        e.n_serie,
                        CASE
                            WHEN e.garantia = 1 THEN 'SI'
                            ELSE 'NO'
                        END AS garantia,
                        e.meses_garantia,
                        CASE
                            WHEN e.compra_biomex = 1 THEN 'SI'
                            ELSE 'NO'
                        END AS compra_biomex,
                        e.fecha_compra,
                        e.comentario
                    FROM equipos AS e
                    WHERE id_Equipos =%s
                    ORDER BY e.id_Equipos DESC
                    """)
                cursor.execute(querySQL, (idEquipo,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_equipoBD: {e}")
        return None
    
# Detalles del SERVICIOS realizado en un equipo
def sql_detalles_servicioBD(idServicio):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_Servicios,
                        e.id_Equipo,
                        CASE
                            WHEN e.tipo_servicio = 1 THEN 'PUESTA EN MARCHA'
                            WHEN e.tipo_servicio = 2 THEN 'LIMPIEZA'
                            WHEN e.tipo_servicio = 3 THEN 'REPARACIÓN'
                            WHEN e.tipo_servicio = 4 THEN 'OTRO SERVICIO'
                            ELSE 'NO ESPECIFICADO'
                        END AS tipo_servicio,
                        e.fecha_servicio,
                        e.trabajador,
                        e.horas,
                        e.comentario
                    FROM servicios AS e
                    WHERE id_Servicios =%s
                    ORDER BY e.id_Servicios DESC
                    """)
                cursor.execute(querySQL, (idServicio,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_servicioBD: {e}")
        return None

# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "Sexo",
                     "Telefono", "Email", "Profesión", "Salario", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        profesion_empleado = registro['profesion_empleado']
        salario_empleado = registro['salario_empleado']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado,
                     salario_empleado, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

# Buscar CLIENTE 
def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_Clientes,
                            e.nombre, 
                            e.apellido,
                            e.nif,
                            e.provincia,
                            e.poblacion,
                            e.direccion,
                            e.codigo_postal,
                            e.telefono,
                            e.email,
                            DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                        FROM clientes AS e
                        WHERE e.nombre LIKE %s 
                        ORDER BY e.id_Clientes DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []

# Buscar EQUIPO 
def buscarEquipoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []

# Buscar CLIENTE unico por ID
def buscarClienteUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_Clientes,
                            e.nombre, 
                            e.apellido,
                            e.nif,
                            e.provincia,
                            e.poblacion,
                            e.direccion,
                            e.codigo_postal,
                            e.telefono,
                            e.email,
                            e.fecha_registro
                        FROM clientes AS e
                        WHERE e.id_Clientes =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarClienteUnico: {e}")
        return []

#Actualizar los datos del CLIENTE
def procesar_actualizacion_form_cliente(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Extraer y procesar datos del formulario
                nombre_cliente = data.form['nombre_cliente']
                apellido_cliente = data.form['apellido_cliente']
                nif_cliente = data.form['nif_cliente']
                telefono_cliente = data.form['telefono_cliente']
                email_cliente = data.form['email_cliente']
                provincia_cliente = data.form['provincia_cliente']
                poblacion_cliente = data.form['poblacion_cliente']
                direccion_cliente = data.form['direccion_cliente']
                cp_cliente = data.form['cp_cliente']

                # Procesar salario eliminando caracteres no numéricos
                id_Clientes = data.form['id_Clientes']

                # Construir consulta SQL y parámetros dinámicamente
                query_base = """
                    UPDATE clientes
                    SET 
                        nombre = %s,
                        apellido = %s,
                        nif = %s,
                        telefono = %s,
                        email = %s,
                        provincia = %s,
                        poblacion = %s,
                        direccion = %s,
                        codigo_postal = %s
                """
                params = [
                    nombre_cliente, apellido_cliente, nif_cliente,
                    telefono_cliente, email_cliente, provincia_cliente, poblacion_cliente,
                    direccion_cliente,cp_cliente
                ]

                # Agregar condición WHERE
                query_base += " WHERE id_Clientes = %s"
                params.append(id_Clientes)

                # Ejecutar la consulta
                cursor.execute(query_base, params)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_cliente: {e}")
        return None

# Buscar EQUIPO unico por ID
def buscarEquipoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_Equipos,
                            e.id_Clientes,
                            e.marca,
                            e.modelo,
                            e.n_serie,
                            e.garantia,
                            e.meses_garantia,
                            e.compra_biomex,
                            e.fecha_compra,
                            e.comentario
                        FROM equipos AS e
                        WHERE e.id_Equipos =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                equipo = mycursor.fetchone()
                return equipo

    except Exception as e:
        print(f"Ocurrió un error en def buscarEquipoUnico: {e}")
        return []
    
#Actualizar los datos del EQUIPO
def procesar_actualizacion_form_equipo(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Extraer y procesar datos del formulario
                marca_equipo = data.form['marca_equipo']
                modelo_equipo = data.form['modelo_equipo']
                ns_equipo = data.form['ns_equipo']
                garantia_equipo = data.form['garantia_equipo']
                mesesgarantia_equipo = data.form['mesesgarantia_equipo']
                compra_equipo = data.form['compra_equipo']
                fecha_equipo = data.form['fecha_equipo']
                comentario_equipo = data.form['comentario_equipo']

                # Procesar salario eliminando caracteres no numéricos
                id_Equipos = data.form['id_Equipos']

                # Construir consulta SQL y parámetros dinámicamente
                query_base = """
                    UPDATE equipos
                    SET 
                        marca = %s,
                        modelo = %s,
                        n_serie = %s,
                        garantia = %s,
                        meses_garantia = %s,
                        compra_biomex = %s,
                        fecha_compra = %s,
                        comentario = %s
                """
                params = [
                    marca_equipo, modelo_equipo, ns_equipo,
                    garantia_equipo, mesesgarantia_equipo, compra_equipo, fecha_equipo,
                    comentario_equipo
                ]

                # Agregar condición WHERE
                query_base += " WHERE id_Equipos = %s"
                params.append(id_Equipos)

                # Ejecutar la consulta
                cursor.execute(query_base, params)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_equipo: {e}")
        return None

# Buscar SERVICIO unico por ID
def buscarServicioUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_Servicios,
                            e.id_Equipo,
                            e.tipo_servicio,
                            e.fecha_servicio,
                            e.trabajador,
                            e.horas,
                            e.comentario
                        FROM servicios AS e
                        WHERE e.id_Servicios =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                servicio = mycursor.fetchone()
                return servicio
    except Exception as e:
        print(f"Ocurrió un error en def buscarServicioUnico: {e}")
        return []

 #Actualizar los datos del SERVICIO
def procesar_actualizacion_form_servicio(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Extraer y procesar datos del formulario
                tipos_servicio = data.form['tipos_servicio']
                fecha = data.form['fecha']
                trabajador_servicio = data.form['trabajador_servicio']
                horas_servicio = data.form['horas_servicio']
                comentario_servicio = data.form['comentario_servicio']

                # Procesar salario eliminando caracteres no numéricos
                id_Servicios = data.form['id_Servicios']

                # Construir consulta SQL y parámetros dinámicamente
                query_base = """
                    UPDATE servicios
                    SET 
                        tipo_servicio = %s,
                        fecha_servicio = %s,
                        trabajador = %s,
                        horas = %s,
                        comentario = %s
                """
                params = [
                    tipos_servicio, fecha, trabajador_servicio, 
                    horas_servicio, comentario_servicio
                ]

                # Agregar condición WHERE
                query_base += " WHERE id_Servicios = %s"
                params.append(id_Servicios)

                # Ejecutar la consulta
                cursor.execute(query_base, params)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_servicio: {e}")
        return None  


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar CLIENTE
def eliminarEmpleado(id_Clientes):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM clientes WHERE id_Clientes=%s"
                cursor.execute(querySQL, (id_Clientes,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
                
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []

# Eliminar EQUIPO
def eliminarEquipo(id_Equipos):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM equipos WHERE id_Equipos=%s"
                cursor.execute(querySQL, (id_Equipos,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
                
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEquipo : {e}")
        return []
    
# Eliminar SERVICIO
def eliminarServicio(id_Servicios):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM servicios WHERE id_Servicios=%s"
                cursor.execute(querySQL, (id_Servicios,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
                
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarServicio : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
    
#Filtrar CLLIENTE por nombre
def obtener_nombres_filtro():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT TRIM(LOWER(nombre)) AS nombre FROM clientes"
                cursor.execute(querySQL,)
                resultados = cursor.fetchall()
                # Capitalizar cada localidad para mostrarla bonita
                filtrar_nombres = [fila['nombre'].capitalize() for fila in resultados]
        return filtrar_nombres

    except Exception as e:
        print(f"Error en obtener_nombres_filtro : {e}")
        return []
#Filtrar CLIENTE por apellido
def obtener_apellidos_filtro():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT TRIM(LOWER(apellido)) AS apellido FROM clientes"
                cursor.execute(querySQL,)
                resultados = cursor.fetchall()
                filtrar_apellidos = [fila['apellido'].capitalize() for fila in resultados]
        return filtrar_apellidos

    except Exception as e:
        print(f"Error en obtener_apellidos_filtro : {e}")
        return []
#Filtrar CLIENTE por población
def obtener_poblaciones_filtro():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT TRIM(LOWER(poblacion)) AS poblacion FROM clientes"
                cursor.execute(querySQL,)
                resultados = cursor.fetchall()
                filtrar_poblaciones = [fila['poblacion'].capitalize() for fila in resultados]
        return filtrar_poblaciones

    except Exception as e:
        print(f"Error en obtener_poblaciones_filtro : {e}")
        return []
    
#Filtro EQUIPO por marca por población
def obtener_marca_filtro():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT TRIM(LOWER(marca)) AS marca FROM equipos"
                cursor.execute(querySQL,)
                resultados = cursor.fetchall()
                filtrar_marca = [fila['marca'].capitalize() for fila in resultados]
        return filtrar_marca

    except Exception as e:
        print(f"Error en obtener_marca_filtro : {e}")
        return []
    
#Filtro EQUIPO por garantia por población
def obtener_garantia_filtro():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT garantia FROM equipos"
                cursor.execute(querySQL)
                resultados = cursor.fetchall()
                # Mapeamos 1 → "Sí", 2 → "No"
                etiqueta_garantia = {
                    1: "SI",
                    2: "NO"
                }
                filtrar_garantia = [etiqueta_garantia[int(fila["garantia"])] for fila in resultados if int(fila["garantia"]) in etiqueta_garantia]
        return filtrar_garantia
    except Exception as e:
        print(f"Error en obtener_garantia_filtro : {e}")
        return []

#Fitro EQUIPO por año de compra
def obtener_filtro_anios():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT YEAR(fecha_compra) AS anio FROM equipos WHERE fecha_compra IS NOT NULL"
                cursor.execute(querySQL)
                resultados = cursor.fetchall()
                anios = [str(fila["anio"]) for fila in resultados if fila["anio"]]
                return sorted(anios)
    except Exception as e:
        print(f"Error en obtener_anios_filtro: {e}")
        return []

#Filtro SERVICIO por tipo de servicio
def obtener_servicio_filtro():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT tipo_servicio FROM servicios"
                cursor.execute(querySQL)
                resultados = cursor.fetchall()
                etiqueta_garantia = {
                    1: "PUESTA EN MARCHA",
                    2: "LIMPIEZA",
                    3: "REPARACIÓN",
                    4: "OTRO SERVICIO"
                }
                filtrar_servicio = [etiqueta_garantia[int(fila["tipo_servicio"])] for fila in resultados if int(fila["tipo_servicio"]) in etiqueta_garantia]
                return filtrar_servicio
    except Exception as e:
        print(f"Error en obtener_servicio_filtro: {e}")
        return []

#Filtro SERVICIO por fecha de servicio
def obtener_filtro_anios_servicio():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT YEAR(fecha_servicio) AS anio_ser FROM servicios WHERE fecha_servicio IS NOT NULL"
                cursor.execute(querySQL)
                resultados = cursor.fetchall()
                anios_ser = [str(fila["anio_ser"]) for fila in resultados if fila["anio_ser"]]
                return sorted(anios_ser)
    except Exception as e:
        print(f"Error en obtener_filtro_anios_servicio: {e}")
        return []
    
#Filtro SERVICIO por mes de servicio
def obtener_filtro_mes_servicio():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT DISTINCT MONTH(fecha_servicio) AS mes_ser FROM servicios WHERE fecha_servicio IS NOT NULL"
                cursor.execute(querySQL)
                resultados = cursor.fetchall()
                meses_ser = [f"{fila['mes_ser']:02d}" for fila in resultados if fila["mes_ser"]]
                return sorted(meses_ser)
    except Exception as e:
        print(f"Error en obtener_filtro_mes_servicio: {e}")
        return []
      
#CLIENTES con equipos
def sql_lista_empleados_con_equipos():
    empleados = sql_lista_empleadosBD()
    for empleado in empleados:
        equipos = sql_lista_equipo_clienteBD(empleado['id_Clientes'])
        # Para cada equipo, añade la lista de servicios
        for equipo in equipos:
            equipo['servicios'] = sql_lista_equipo_servicioBD(equipo['id_Equipos'])
        empleado['equipos'] = equipos
    return empleados

#Busqueda avanzada a partir de filtros
def buscarAvanzadoBD(nombre='', apellido='', poblacion='', marca='', garantia='', anio='', tipo_servicio='', anio_servicio='', mes_servicio=''):
    try:
        print("🟡 Filtros recibidos:")
        print(f"Nombre: {nombre}")
        print(f"Apellido: {apellido}")
        print(f"Población: {poblacion}")
        print(f"Marca: {marca}")
        print(f"Garantía: {garantia}")
        print(f"Año de compra: {anio}")
        print(f"Tipo de servicio: {tipo_servicio}")
        print(f"Año de servicio: {anio_servicio}")
        print(f"Mes de servicio: {mes_servicio}")

        condiciones = []
        valores = []

        # Filtros por CLIENTE
        if nombre:
            condiciones.append("TRIM(LOWER(nombre)) = %s")
            valores.append(nombre.strip().lower())
        if apellido:
            condiciones.append("TRIM(LOWER(apellido)) = %s")
            valores.append(apellido.strip().lower())
        if poblacion:
            condiciones.append("TRIM(LOWER(poblacion)) = %s")
            valores.append(poblacion.strip().lower())

        query = "SELECT * FROM clientes"
        if condiciones:
            query += " WHERE " + " AND ".join(condiciones)

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                cursor.execute(query, valores)
                clientes = cursor.fetchall()

        resultado = []
        for cliente in clientes:
            equipos = sql_lista_equipo_clienteBD(cliente['id_Clientes'])
            equipos_filtrados = []

            for equipo in equipos:
                # Filtrar por marca
                if marca and equipo.get("marca", "").strip().lower() != marca.strip().lower():
                    continue

                # Filtrar por garantía
                if garantia and equipo.get("garantia", "").strip().lower() != garantia.strip().lower():
                    continue

                # Filtrar por año de compra
                if anio:
                    fecha_compra = equipo.get("fecha_compra", "")
                    if not fecha_compra or not fecha_compra.startswith(str(anio)):
                        continue

                # Obtener servicios asociados al equipo
                servicios = sql_lista_equipo_servicioBD(equipo["id_Equipos"])
                servicios_filtrados = []

                for servicio in servicios:
                    # Filtro por tipo de servicio
                    if tipo_servicio and servicio.get("tipo_servicio", "").strip().lower() != tipo_servicio.strip().lower():
                        continue

                    # Filtro por año de servicio
                    if anio_servicio:
                        fecha_servicio = servicio.get("fecha_servicio", "")
                        if not fecha_servicio or not fecha_servicio.startswith(str(anio_servicio)):
                            continue

                    # Filtro por mes de servicio (mes con dos dígitos)
                    if mes_servicio:
                        fecha_servicio = servicio.get("fecha_servicio", "")
                        if not fecha_servicio:
                            continue
                        # Extraer el mes en formato MM
                        mes = fecha_servicio[5:7]
                        if mes != mes_servicio:
                            continue

                    servicios_filtrados.append(servicio)

                # Si hay filtros de servicios aplicados y no quedan servicios válidos, excluir el equipo
                if (tipo_servicio or anio_servicio or mes_servicio) and not servicios_filtrados:
                    continue

                equipo["servicios"] = servicios_filtrados if (tipo_servicio or anio_servicio or mes_servicio) else servicios
                equipos_filtrados.append(equipo)

            # Si aplicamos cualquier filtro de equipo o servicio y no quedan equipos, excluimos el cliente
            if (marca or garantia or anio or tipo_servicio or anio_servicio or mes_servicio) and not equipos_filtrados:
                continue

            cliente["equipos"] = equipos_filtrados if (marca or garantia or anio or tipo_servicio or anio_servicio or mes_servicio) else equipos
            resultado.append(cliente)

        return resultado

    except Exception as e:
        print(f"Error en buscarAvanzadoBD: {e}")
        return []

